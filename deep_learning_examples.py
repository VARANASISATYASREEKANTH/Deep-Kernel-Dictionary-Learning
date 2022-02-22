# -*- coding: utf-8 -*-
"""
Created on Fri Jul 23 13:46:00 2021

@author: asdg
"""
# -*- coding: utf-8 -*-
"""
Created on Thu Jul  8 09:49:42 2021

@author: asdg
"""
#------for convolution layer-----
import keras
from keras.models import Sequential,Input,Model
from keras.layers import Dense, Dropout, Flatten
from keras.layers import Conv2D, MaxPooling2D
from keras.layers.normalization import BatchNormalization
from keras.layers.advanced_activations import LeakyReLU
#----------for LSTM-----------
import tensorflow as tf
from numpy import array
from tensorflow.keras.models import Model
from keras.layers import Input
from keras.layers import LSTM
from keras.layers import Dense
from keras.layers import RepeatVector
from keras.layers import TimeDistributed
from keras.utils import plot_model
import keras
import xlrd
import numpy as np
import matplotlib.pyplot as plt
import math
from random import *
from math import log 
from scipy.signal import savgol_filter
import matplotlib.pyplot as plt
import keras
from keras.models import Sequential,Input,Model
from keras.layers import Dense, Dropout, Flatten
from keras.layers import Conv2D, MaxPooling2D
from keras.layers.normalization import BatchNormalization
from keras.layers.advanced_activations import LeakyReLU
from sklearn import preprocessing
import openpyxl
from pathlib import Path


plt.rcParams.update({'font.size':10})
params = {'backend': 'ps',
          'axes.labelsize': 10,
          'legend.fontsize':10,
          'xtick.labelsize':10,
          'ytick.labelsize':10,
          'text.usetex': True};
#--------------80% train data, 20% test data----------

loc1=('G:/research_works_vssreekanth_jrf/MY_PAPERS/paper_3a_deep_learning_for_parameter_estimation/programs/anamoly_detection_LSTM_RNN/training_data/train_data_temperature_perturbations_ver2.xlsx')
wb = openpyxl.load_workbook(loc1)
sheet = wb.active

T_pert=np.zeros((sheet.max_row,sheet.max_column))

for i in range(1,1,124):
    for j in range(1,1,1024):
        k=sheet.cell(row=i,column=j);
        print(k);
        T_pert[i][j]=k.value;
#-----        
for i in range(1,1,124):
    for j in range(1,1,1024):
        T_pert[i][j]=T_pert[i][j]-np.mean(T_pert[i][:])
        
    
         
        
#------------creating a tensor of temperature perturbations        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        